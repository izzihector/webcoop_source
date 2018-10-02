# -*- coding: utf-8 -*-
###############################################
# EzTech Software & Consultancy Inc. (c) 2017
###############################################

from odoo import api, fields, models
from odoo import tools, _
from odoo.exceptions import ValidationError, Warning
from odoo.modules.module import get_module_resource
from datetime import datetime
from dateutil.relativedelta import relativedelta
from odoo.tools.safe_eval import safe_eval
import logging
from openpyxl import load_workbook
import base64
from copy import copy
try:
    from cStringIO import StringIO
except ImportError:
    from StringIO import StringIO

_logger = logging.getLogger(__name__)
DF = "%Y-%m-%d"

def set_cell_format(new_cell, ref_cell):
    new_cell.font = copy(ref_cell.font)
    new_cell.border = copy(ref_cell.border)
    new_cell.fill = copy(ref_cell.fill)
    new_cell.number_format = copy(ref_cell.number_format)
    new_cell.alignment = copy(ref_cell.alignment)

def set_cell(ws_out, ws_ref, ccol, row, value):
    cname = ccol + ("%d" % row)
    cref = '%s15' % ccol
    new_cell = ws_out[cname]
    ref_cell = ws_ref[cref]
    new_cell.value = value
    set_cell_format(new_cell, ref_cell)

class ExcelReportWizard(models.TransientModel):
    _name = 'wc.excel.report.wizard'
    _description = 'Wizard for Excel Reports'

    #default-last closed date
    date1 = fields.Date("Date From", required=True, default=fields.Date.context_today)
    date2 = fields.Date("Date To", required=True, default=fields.Date.context_today)
    report_id = fields.Many2one('wc.excel.report', required=True, ondelete='cascade', string='Report')

    prepared_by = fields.Char("Prepared By")
    checked_by = fields.Char("Checked By")
    approved_by = fields.Char("Approved By")

    journal_ids = fields.Many2many('account.journal', string='Journals')
    account_ids = fields.Many2many('account.account', string='Accounts')
    voucher_ref = fields.Char("Voucher Ref")

    partner_id = fields.Many2one('res.partner', string='Partner')

    #report_type = fields.Selection([
    #    ('daily-cash-position', 'Daily Cash Position'),
    #], 'Report Type', required=True, default="daily-cash-position")

    filename = fields.Char("Filename", default="output.xlsx")
    excel_file = fields.Binary("Excel File", readonly=True)

    @api.onchange('report_id')
    def onchange_report_id(self):
       self.ensure_one()
       self.prepared_by = self.report_id.prepared_by
       self.checked_by = self.report_id.checked_by
       self.approved_by = self.report_id.approved_by
       self.journal_ids = self.report_id.journal_ids
       self.account_ids = self.report_id.account_ids

    @api.model
    def get_account_name(self, code):
        scode = ("%s" % code).strip()
        res = self.env['account.account'].search([
            ('code','=',scode)
        ], limit=1)
        if res:
            return res[0].name
        else:
            return "Account code not found!"

    def get_acct_totals(self, code1, code2, where_str="", where_params=[]):
        scode1 = ("%s" % code1).strip()
        scode2 = ("%s" % code2).strip()
        accounts = self.env['account.account'].search([
            ('code','>=',scode1),
            ('code','<=',scode2)
        ])

        if not accounts:
            return {}

        if where_str:
            where_str = " AND " + where_str

        request = (
            "SELECT SUM(debit) AS debit," +
            " SUM(credit) AS credit," +
            " (SUM(debit) - SUM(credit)) AS balance" +
            " FROM account_move_line AS line" +
            " INNER JOIN account_move AS move ON move.id=line.move_id" +
            " WHERE line.account_id IN %s" +
            " AND move.state='posted'" +
            where_str
        )
        params = (tuple(accounts.ids),) + tuple(where_params)
        _logger.debug("sql request: %s param=%s", request, params)
        self.env.cr.execute(request, params)
        for row in self.env.cr.dictfetchall():
            return row


    @api.model
    def gen_excel_jsummary(self, wiz, template):

        template_io = StringIO(template.decode('base64'))
        wb = load_workbook(template_io)
        if 'data' not in wb.get_sheet_names():
            raise Warning(_("Worksheet 'data' does not exist."))

        date1 = fields.Date.from_string(wiz.date1)
        date2 = fields.Date.from_string(wiz.date2)
        now = fields.Datetime.context_timestamp(self, datetime.now())

        ws = wb['data']
        ws = wb['data']
        ws['B1'] = date1.strftime("%B %d, %Y")
        ws['B2'] = date2.strftime("%B %d, %Y")
        ws['B3'] = now.strftime("%A, %d. %B %Y %I:%M%p")
        ws['B6'] = wiz.prepared_by or " "
        ws['B7'] = wiz.checked_by or " "
        ws['B8'] = wiz.approved_by or " "

        start_row =  int(ws['B12'].value)

        sql = """
            SELECT l.account_id as account_id,
                a.code AS code,
                a.name AS name,
                SUM(l.debit) AS tdebit,
                SUM(l.credit) AS tcredit
            FROM account_move_line AS l
            INNER JOIN account_move AS m ON m.id = l.move_id
            INNER JOIN account_account AS a ON a.id = l.account_id
            WHERE m.state='posted' AND %s
            GROUP BY 1,2,3
            ORDER BY 2,3
        """

        where_str = "l.company_id=%s AND l.date>=%s AND l.date<=%s"
        params = [wiz.report_id.company_id.id, date1, date2]
        if wiz.journal_ids:
            wstr = " AND l.journal_id IN %s"
            ids = tuple(wiz.journal_ids.ids)
            _logger.debug("Journal filter: %s %s", wstr, ids)
            where_str += wstr
            params += [ids]
        if wiz.account_ids:
            wstr = " AND l.account_id IN %s"
            ids = tuple(wiz.account_ids.ids)
            _logger.debug("Account filter: %s %s", wstr, ids)
            where_str += wstr
            params += [ids]
        if wiz.partner_id:
            wstr = " AND l.partner_id = %s"
            _logger.debug("Partner filter: %s %s", wstr, wiz.partner_id.id)
            where_str += wstr
            params += [wiz.partner_id.id]

        request = sql % where_str
        _logger.debug("sql request: %s param=%s", request, params)
        self.env.cr.execute(request, params)

        ws0 = wb.worksheets[0]
        tdebit = 0.0
        tcredit = 0.0
        for rec in self.env.cr.dictfetchall():
            _logger.debug("data: %s", rec)
            debit = rec['tdebit']
            credit = rec['tcredit']
            if debit or credit:
                #account_id = self.env['account.account'].browse(rec['account_id'])
                set_cell(ws0, ws, 'A', start_row, rec['code'])
                set_cell(ws0, ws, 'B', start_row, rec['name'])
                set_cell(ws0, ws, 'C', start_row, debit)
                set_cell(ws0, ws, 'D', start_row, credit)

                start_row += 1
                tdebit += debit
                tcredit += credit

        ws['B4'].value = tdebit
        ws['B5'].value = tcredit

        #footer
        start_row += 2
        for i in range(0,3):
            for c in ['A','B','C','D']:
                ncell = ws0['%s%d' % (c, start_row+i)]
                rcell = ws['%s%d' % (c, i+20)]
                ncell.value = rcell.value
                set_cell_format(ncell, rcell)

        output = StringIO()
        wb.save(output)
        output_value = output.getvalue()
        res = base64.encodestring(output_value)
        output.close()
        return res

    @api.model
    def gen_excel_summary(self, wiz, template):
        START_DATA_ROW = 11
        #ws.delete_rows(index, 1)

        template_io = StringIO(template.decode('base64'))
        wb = load_workbook(template_io)
        if 'data' not in wb.get_sheet_names():
            raise Warning(_("Worksheet 'data' does not exist."))

        date1 = fields.Date.from_string(wiz.date1)
        date2 = fields.Date.from_string(wiz.date2)
        now = fields.Datetime.context_timestamp(self, datetime.now())

        ws = wb['data']
        ws['B1'] = date1.strftime("%B %d, %Y")
        ws['B2'] = date2.strftime("%B %d, %Y")
        ws['B3'] = now.strftime("%A, %d. %B %Y %I:%M%p")

        ws['B6'] = wiz.prepared_by or " "
        ws['B7'] = wiz.checked_by or " "
        ws['B8'] = wiz.approved_by or " "

        for r in range(START_DATA_ROW, ws.max_row+1): #ws.iter_cols(min_row=3, min_col=1, max_col=2):
            acct_from = ws['A%d' % r].value
            acct_to = ws['B%d' % r].value
            _logger.debug("row %d: %s-%s",r,acct_from,acct_to)

            if acct_from and acct_to:
                ws['C%d' % r] = self.get_account_name(acct_from)
                ws['D%d' % r] = self.get_account_name(acct_to)

                bbal = self.get_acct_totals(acct_from, acct_to,
                    where_str="line.date<%s",
                    where_params=[date1.strftime(DF)])
                period = self.get_acct_totals(acct_from, acct_to,
                    where_str="line.date>=%s AND line.date<=%s",
                    where_params=[date1.strftime(DF), date2.strftime(DF)])
                _logger.debug("bbal: %s",bbal)

                ws['E%d' % r] = bbal['balance']
                ws['F%d' % r] = period['debit']
                ws['G%d' % r] = period['credit']
                ws['H%d' % r].value = "=F%d-G%d" % (r,r)
                ws['I%d' % r].value = "=E%d+F%d-G%d" % (r,r,r)

        output = StringIO()
        wb.save(output)
        output_value = output.getvalue()
        res = base64.encodestring(output_value)
        output.close()

    ##################################
    @api.model
    def get_sql_filter(self, where_str="",params=[]):
        if self.journal_ids:
            wstr = " AND m.journal_id IN %s"
            ids = tuple(self.journal_ids.ids)
            _logger.debug("Journal filter: %s %s", wstr, ids)
            where_str += wstr
            params += [ids]
        if self.account_ids:
            wstr = " AND l.account_id IN %s"
            ids = tuple(self.account_ids.ids)
            _logger.debug("Account filter: %s %s", wstr, ids)
            where_str += wstr
            params += [ids]
        if self.partner_id:
            wstr = " AND l.partner_id = %s"
            pid = self.partner_id.id
            _logger.debug("Partner filter: %s %s", wstr, pid)
            where_str += wstr
            params += [pid]
        return where_str, params

    @api.model
    def get_data_details(self, add_where_str="", add_params=[], order="2,1"):

        date1 = fields.Date.from_string(self.date1)
        date2 = fields.Date.from_string(self.date2)

        sql = """
            SELECT
                l.id AS id,
                l.date AS date,
                j.name AS journal,
                m.name AS voucher,
                m.ref AS ref,
                a.code AS code,
                a.name AS account,
                l.name AS label,
                p.name AS partner,
                l.debit AS debit,
                l.credit AS credit
            FROM account_move_line AS l
            INNER JOIN account_move AS m ON m.id = l.move_id
            INNER JOIN account_account AS a ON a.id = l.account_id
            INNER JOIN account_journal AS j ON j.id = m.journal_id
            LEFT JOIN res_partner AS p ON p.id = l.partner_id
            WHERE m.state='posted' $where_str
            ORDER BY $order_str
            """

        params = [self.report_id.company_id.id, date1, date2]
        where_str = " AND l.company_id=%s AND l.date>=%s AND l.date<=%s"

        where_str += add_where_str
        params += add_params
        where_str, params = self.get_sql_filter(where_str, params)

        request = sql.replace("$where_str", where_str).replace("$order_str", order)
        _logger.debug("sql request: %s param=%s", request, params)

        self.env.cr.execute(request, params)
        return self.env.cr.dictfetchall()

    @api.model
    def get_data_lines(self, add_where_str="", add_params=[], totals=False, order="2,3"):

        date1 = fields.Date.from_string(self.date1)
        date2 = fields.Date.from_string(self.date2)

        if totals:
            sql = """
                SELECT
                    SUM(CASE WHEN l.date<%s THEN l.debit END) as bdebit,
                    SUM(CASE WHEN l.date<%s THEN l.credit END) as bcredit,
                    SUM(CASE WHEN l.date>=%s AND l.date<=%s THEN l.debit END) as debit,
                    SUM(CASE WHEN l.date>=%s AND l.date<=%s THEN l.credit END) as credit,
                    SUM(l.debit) AS tdebit,
                    SUM(l.credit) AS tcredit
                FROM account_move_line AS l
                INNER JOIN account_move AS m ON m.id = l.move_id
                INNER JOIN account_account AS a ON a.id = l.account_id
                WHERE m.state='posted' $where_str
            """

        else:
            sql = """
                SELECT l.account_id as account_id,
                    a.code AS code,
                    a.name AS name,
                    SUM(CASE WHEN l.date<%s THEN l.debit END) as bdebit,
                    SUM(CASE WHEN l.date<%s THEN l.credit END) as bcredit,
                    SUM(CASE WHEN l.date>=%s AND l.date<=%s THEN l.debit END) as debit,
                    SUM(CASE WHEN l.date>=%s AND l.date<=%s THEN l.credit END) as credit,
                    SUM(l.debit) AS tdebit,
                    SUM(l.credit) AS tcredit
                FROM account_move_line AS l
                INNER JOIN account_move AS m ON m.id = l.move_id
                INNER JOIN account_account AS a ON a.id = l.account_id
                WHERE m.state='posted' $where_str
                GROUP BY 1,2,3
                ORDER BY $order_str
            """

        params = [date1, date1, date1, date2, date1, date2]
        where_str = " AND l.company_id=%s"
        params += [self.report_id.company_id.id]
        where_str += " AND l.date<=%s"
        params += [date2]

        where_str += add_where_str
        params += add_params
        where_str, params = self.get_sql_filter(where_str, params)

        request = sql.replace("$where_str", where_str).replace("$order_str", order)
        _logger.debug("sql request: %s param=%s", request, params)

        self.env.cr.execute(request, params)
        return self.env.cr.dictfetchall()

    @api.model
    def gen_excel_jsummary2(self, wiz, template):

        template_io = StringIO(template.decode('base64'))
        wb = load_workbook(template_io)
        if 'data' not in wb.get_sheet_names():
            raise Warning(_("Worksheet 'data' does not exist."))

        date1 = fields.Date.from_string(wiz.date1)
        date2 = fields.Date.from_string(wiz.date2)
        now = fields.Datetime.context_timestamp(self, datetime.now())

        wsdata = wb['data']
        wsformat = wb['format']
        wsout = wb.worksheets[0]

        ws = wb['data']
        ws['B1'] = date1.strftime("%B %d, %Y")
        ws['B2'] = date2.strftime("%B %d, %Y")
        ws['B3'] = now.strftime("%A, %d. %B %Y %I:%M%p")
        ws['B6'] = wiz.prepared_by or " "
        ws['B7'] = wiz.checked_by or " "
        ws['B8'] = wiz.approved_by or " "

        fill_row =  int(ws['B9'].value)
        col_count = int(ws['B10'].value)
        act_row = fill_row
        irows = {}

        while True:
            filter = wsformat['A%d' % act_row].value or False
            _logger.debug("filter %d: filter=%s", act_row, filter)
            if (not filter) or filter[0]!='$':
                break

            xfilter = filter.split(",")
            _logger.debug("cmd %s: %s", act_row, xfilter)

            cmd = xfilter[0]

            if cmd in ['$total', '$acct', '$detail']:
                if len(xfilter)<4:
                    act_row += 1
                    continue

                acct1 = str(xfilter[1])
                acct2 = str(xfilter[2])
                skip_zero = int(xfilter[3] or 0)

                where_str = " AND a.code>=%s AND a.code<=%s"
                params = [acct1, acct2]

                if wiz.voucher_ref:
                    where_str += " AND m.ref ILIKE %s"
                    params.append("%%%s%%" % wiz.voucher_ref.strip())

                if cmd == '$total':
                    #data_lines = [wiz.get_acct_totals(acct1, acct2)]
                    data_lines = wiz.get_data_lines(where_str, params, totals=True)
                elif cmd=='$detail':
                    data_lines = wiz.get_data_details(where_str, params)
                else:
                    #$acct
                    data_lines = wiz.get_data_lines(where_str, params)

                if not data_lines:
                    data_lines = [{}]

            #$in,acct,acct,...,skipzero
            elif cmd=='$in':
                if len(xfilter)<3:
                    act_row += 1
                    continue

                skip_zero = int(xfilter[-1] or 0)
                where_str = " AND a.code IN %s"
                params = [tuple(xfilter[1:-1])]
                data_lines = wiz.get_data_lines(where_str, params, totals=True)

            elif cmd=='$copy':
                data_lines = [{}]
                skip_zero = 0
            else:
                act_row += 1
                continue

            totals = {}
            for rec in data_lines:

                if act_row not in irows:
                    irows[act_row] = [fill_row, fill_row]

                _logger.debug("data r=%s: %s irows=%s", fill_row, rec, irows)
                debit = rec.get('debit')
                credit = rec.get('credit')
                code_name = "%s - %s" % (rec.get('code'), rec.get('name'))
                for k in ['debit','credit','bdebit','bcredit','tdebit','tcredit']:
                    if not rec.get(k):
                        rec[k] = 0.0
                    totals[k] = totals.get(k, 0.0) + rec[k]

                local_dict = rec.copy()
                local_dict['totals'] = totals
                local_dict['code_name'] = code_name

                # 1-debit, 2-credit, 3-both, 0-noskip, 4-total

                if skip_zero==1 and debit:
                    is_skip = False
                elif skip_zero==2 and credit:
                    is_skip = False
                elif skip_zero==3 and (debit or credit):
                    is_skip = False
                elif skip_zero==0:
                    is_skip = False
                else:
                    is_skip = True

                #_logger.debug("skipz %s: is_skip=%d d=%s c=%s", skip_zero, is_skip, debit, credit)
                #if (not skip_zero) or debit or credit:

                if not is_skip:

                    irows[act_row][1] = fill_row

                    for c in range(col_count):
                        data_key = "%s%d" % (chr(65+c), fill_row)
                        ref_key = "%s%d" % (chr(66+c), act_row)
                        newcell = wsout[data_key]
                        refcell = wsformat[ref_key]
                        v = refcell.value
                        _logger.debug("v %s: %s", ref_key, v)
                        if v:
                            if v[:4]=='$sum':
                                xv = v.split(',')
                                r1 = irows[int(xv[2])][0]
                                r2 = irows[int(xv[3])][1]
                                if r2==fill_row:
                                    r2 -= 1
                                formula = "=SUM(%s%s:%s%s)" % (
                                    xv[1],
                                    r1,
                                    xv[1],
                                    r2,
                                )
                                _logger.debug("sum r=%s: %s", fill_row, formula)
                                newcell.value = formula
                            elif v[:4]=='$add':
                                xv = v.split(',')
                                _logger.debug("add0 r=%s: xv=%s", xv)
                                formula = "="
                                for i in range(1, len(xv)):
                                    code = xv[i]
                                    if code[0] not in ['+','-']:
                                        code = "+" + code
                                    r0 = int(code[2:])
                                    r = irows[r0][0]
                                    formula += "%s%s" % (code[:2],r)
                                _logger.debug("add r=%s: %s", fill_row, formula)
                                newcell.value = formula
                            elif v[:6]=='$eval:':
                                newcell.value = safe_eval(v[6:], locals_dict=local_dict)
                            else:
                                newcell.value = v
                        set_cell_format(newcell, refcell)

                    fill_row += 1

            act_row += 1

        #hide config sheets
        wsdata.sheet_state = 'hidden'
        wsformat.sheet_state = 'hidden'

        output = StringIO()
        wb.save(output)
        output_value = output.getvalue()
        res = base64.encodestring(output_value)
        output.close()
        return res


    @api.multi
    def gen_report(self):
        self.ensure_one()
        is_generated = False
        if self.report_id and self.report_id.template_data:
            template = self.report_id.template_data
            if template == False:
                raise Warning(_("Excel report template is not defined."))
            self.filename = self.report_id.template_filename

            if self.report_id.report_type=='summary':
                self.excel_file = self.gen_excel_summary(self, template)
                is_generated = True

            elif self.report_id.report_type=='jsummary':
                self.excel_file = self.gen_excel_jsummary2(self, template)
                is_generated = True

        if is_generated:
            return {
                'type': 'ir.actions.report.xml',
                'report_type': 'controller',
                'report_file': '/web/content/wc.excel.report.wizard/%s/excel_file/%s?download=true' % (
                    self.id,
                    self.filename,
                )
            }
        else:
            return {
                'name': 'Excel Report',
                'context': self.env.context,
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'wc.excel.report.wizard',
                'res_id': self.id,
                'view_id': False,
                'type': 'ir.actions.act_window',
                'target': 'new',
            }



#
